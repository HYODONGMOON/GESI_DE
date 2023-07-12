# -*- coding: utf-8 -*-

import tkinter as tk
import subprocess
import tkinter.messagebox as messagebox
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import random
import numpy as np
from openpyxl import load_workbook
from tkinter import ttk


# 전체 인터페이스 코드 #

def handle_menu_selection(event):
    selected_item = menu_var.get()
    print(f"Selected: {selected_item}")

window = tk.Tk()
# window.geometry("700x800")

# 최상단 메뉴 패널
menubar = tk.Menu(window)
window.config(menu=menubar)

# "파일" 메뉴
file_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="파일", menu=file_menu)
file_menu.add_command(label="열기")
file_menu.add_command(label="저장")
file_menu.add_separator()
file_menu.add_command(label="종료")

# "편집" 메뉴
edit_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="편집", menu=edit_menu)
edit_menu.add_command(label="복사")
edit_menu.add_command(label="붙여넣기")
edit_menu.add_separator()
edit_menu.add_command(label="실행 취소")

# "도움말" 메뉴
help_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="도움말", menu=help_menu)
help_menu.add_command(label="도움말 보기")
help_menu.add_command(label="버전 정보")

# 나머지 인터페이스 그리드
# ...


# 모델 분석관련 코드 #

def toggle_entry(entry, var):
    if var.get():
        entry.config(state='normal')
    else:
        entry.config(state='disabled')

# 입력값 제한
def get_cell_value(file_path, sheet_name, cell):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    value = sheet[cell].value
    workbook.close()
    return value

def save_to_excel():
    input_fields = {
        "감축목표": {
            "entry": target_goal_entry, 
            "min_val": 0,
            "max_val": 100
            },
        "에너지 자급률": {
            "entry": energy_self_entry,
            "min_val": 0, 
            "max_val": 50
            },
        "인구 변화": {
            "entry": population_change_entry, 
            "min_val": 0, 
            "max_val": 50
            } 
        }

    error_messages = []

    for field_name, field_info in input_fields.items():
        entry = field_info["entry"]
        value = entry.get()
        min_val = field_info["min_val"]
        max_val = field_info["max_val"]

        if not validate_input(value, min_val, max_val):
            error_messages.append(field_name)
            entry.config(bg="red")  # 잘못된 입력 필드에 빨간색 음영 적용
        else:
            entry.config(bg="white")  # 올바른 입력 필드는 흰색으로 설정

    if error_messages:
        show_error_message(error_messages)
        return

    # 공통 #
    target_year = int(target_year_entry.get())
    target_goal = float(target_goal_entry.get())
    energy_demand_change = float(energy_demand_change_entry.get())
    energy_self = float(energy_self_entry.get())
    land_utilisation = float(land_utilisation_entry.get())

    # 수요도시형 #        
    city_area = float(city_area_entry.get())
    residential_area = float(residential_area_entry.get())
    commercial_area = float(commercial_area_entry.get())
    transportation_facility = float(transportation_facility_entry.get())
    environmental_facility = float(environmental_facility_entry.get())

    population_change = float(population_change_entry.get())
    electricity_share_city = float(electricity_share_city_entry.get())
    h2_share_city = float(h2_share_city_entry.get())
    waste_utilisation = float(waste_utilisation_entry.get())

    # 산업중심형 #
    industry_area = float(regional_area_entry.get())
    iron_area = float(iron_area_entry.get())
    refinery_area = float(refinery_area_entry.get())
    petro_area = float(petro_area_entry.get())
    cement_area = float(cement_area_entry.get())

    eaf = float(eaf_entry.get())
    dri = float(dri_entry.get())
    chemical_re = float(chemical_re_entry.get())
    h2_utilisation = float(h2_utilisation_entry.get())
    fuel_eff = float(fuel_eff_entry.get())
    electricity_share_industry = float(electricity_share_entry.get())

    # 농어촌지역형 #
    rural_area = float(rural_area_entry.get())
    land_rural_area = float(land_rural_area_entry.get())
    sea_rural_area = float(sea_rural_area_entry.get())
    cultivated_area = float(cultivated_area_entry.get())
    forest = float(forest_entry.get())
    fish_farm = float(fish_farm_entry.get())
    fishing_boat = float(fishing_boat_entry.get())

    trees = float(trees_entry.get())
    rice = float(rice_entry.get())
    animals = float(animals_entry.get())
    fishfarm_change = float(fishfarm_change_entry.get())
    electric_boat = float(electric_boat_entry.get())
    h2_boat = float(h2_boat_entry.get())
    P2G_share = float(P2G_share_entry.get())

    # 항만인프라형 #
    port_area = float(port_area_entry.get())
    sea_port_area = float(sea_port_area_entry.get())
    cargo_cap = float(cargo_cap_entry.get())    

    petro_share = float(petro_share_entry.get())
    trans_h2 = float(trans_h2_entry.get())
    liquid_h2 = float(liquid_h2_entry.get())  
    amonia = float(amonia_entry.get())
    storage_cap = float(storage_cap_entry.get())
    sea_area_change = float(sea_area_change_entry.get())  
    density_change = float(density_change_entry.get())
    dedicated_h2 = float(dedicated_h2_entry.get())  


    wb = load_workbook('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')
    ws = wb['Scenario']
    
    # 기본 전제조건
    ws['E4'] = target_year
    ws['E5'] = target_goal /100
    ws['E7'] = energy_demand_change /100
    ws['E11'] = energy_self /100
    ws['E16'] = land_utilisation /100

    #수요도시형 시나리오
    ws['E22'] = population_change /100
    ws['E23'] = population_change /100
    ws['E25'] = electricity_share_city /100
    ws['E28'] = electricity_share_city /100
    ws['E27'] = h2_share_city /100
    ws['E29'] = h2_share_city /100
    ws['E31'] = waste_utilisation /100

    #산업중심형 시나리오
    ws['E40'] = eaf /100
    ws['E41'] = dri /100
    ws['E42'] = chemical_re /100
    ws['E43'] = h2_utilisation /100
    ws['E45'] = fuel_eff /100
    ws['E46'] = electricity_share_industry /100

    #농어촌지역형 시나리오
    ws['E50'] = trees /100
    ws['E51'] = rice /100
    ws['E52'] = animals /100
    ws['E54'] = fishfarm_change /100
    ws['E55'] = electric_boat /100
    ws['E56'] = h2_boat /100
    ws['E58'] = P2G_share /100

    #항만인프라형 시나리오
    ws['E62'] = petro_share /100
    ws['E63'] = trans_h2 /100
    ws['E64'] = liquid_h2 /100
    ws['E65'] = amonia /100
    ws['E66'] = storage_cap 
    ws['E68'] = sea_area_change /100
    ws['E69'] = density_change /100
    ws['E70'] = dedicated_h2 /100


    ws1 = wb['Demand']

    #수요도시형 
    ws1['E4'] = city_area
    ws1['E5'] = residential_area /100
    ws1['E6'] = commercial_area /100
    ws1['E7'] = transportation_facility /100
    ws1['E8'] = environmental_facility /100

    #산업중심형
    ws1['E18'] = industry_area
    ws1['E19'] = iron_area /100
    ws1['E20'] = refinery_area /100
    ws1['E21'] = petro_area /100
    ws1['E22'] = cement_area /100

    #농어촌지역형
    ws1['E31'] = rural_area
    ws1['E32'] = land_rural_area /100
    ws1['E33'] = sea_rural_area /100
    ws1['E35'] = cultivated_area /100
    ws1['E36'] = forest /100
    ws1['E41'] = fish_farm /100
    ws1['E42'] = fishing_boat /100

    #항만인프라형 시나리오
    ws1['E47'] = port_area 
    ws1['E48'] = sea_port_area 
    ws1['E50'] = cargo_cap


    wb.save('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')

   

def run_script():
    subprocess.call(["python", "C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\gesi.py"])

def toggle_scenario():
    if all(region in selected_regions for region in regions):
        scenario_frame_city.grid(row=4, column=0, columnspan=2, padx=5, pady=5, sticky="w")
        scenario_frame_industry.grid(row=4, column=2, columnspan=2, padx=5, pady=5, sticky="e")
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "수요도시형" in selected_regions:
        scenario_frame_city.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "산업중심형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "농어촌지역형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_port.grid_forget()
    elif "항만인프라형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
    else:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()

def validate_input(value, min_val, max_val):
    try:
        num = float(value)
        if min_val <= num <= max_val:
            return True
        else:
            return False
    except ValueError:
        return False

def show_error_message(error_messages):
    error_text = "\n".join(f" - {field_name}" for field_name in error_messages)
    messagebox.showerror("입력 오류", f"다음 입력 필드가 올바른 범위 내에 입력되지 않았습니다:\n{error_text}")


# 데이터 시각화 #


def show_result():
    # 엑셀 파일에서 데이터 읽기
    df = pd.read_excel('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\results_2050.xlsx', sheet_name='Figure_hour_electricity', header=None, usecols="X:AO", nrows=73)

    fig, ax = plt.subplots()

    # X행 선 그래프
    x = [str(i) for i in df.index[1:]]  # 정수형 인덱스를 문자열로 변환
    y = df.iloc[1:, 0].apply(pd.to_numeric)  # X2:X169 셀 값을 Y 데이터로 사용 (부동소수점으로 변환)
    ax.plot(x[:len(y)], y, 'b-', label='Line Graph')  # 선 그래프에 레이블 추가
    ax.set_xlabel('X')
    ax.set_ylabel('Y', color='b')
    ax.tick_params('y', colors='b')

    # Y~AO행 양수 막대 그래프
    positive_data = df.iloc[1:, 1:].apply(pd.to_numeric, errors='coerce')  # 문자열을 부동소수점으로 변환
    num_columns = positive_data.shape[1]  # 데이터 열의 개수
    colors = ['#' + ''.join(random.choice('0123456789ABCDEF') for _ in range(6)) for _ in range(num_columns)]  # 랜덤 색상 생성

    width = 0.5  # 막대 그래프 너비
    positions = np.arange(len(x[:len(y)]))  # 막대 그래프 위치

    column_sums = positive_data.sum(axis=0)
    sorted_indexes = column_sums.argsort()[::-1]  # 열의 전체 합이 큰 순서대로 정렬

    handles, labels = [], []
    for i, col in enumerate(positive_data.columns[sorted_indexes]):
        labels = [str(label) for label in df.iloc[0, 1:].tolist()]  # 0행의 값을 문자열로 변환하여 범례로 사용
        positive_values = positive_data[col][positive_data[col] > 0]
        negative_values = positive_data[col][positive_data[col] < 0]

        if positive_values.sum() != 0:
            handles.append(ax.bar(positions[:len(positive_values)], positive_values, width, color=colors[i], label=labels[i]))

        if negative_values.sum() != 0:
            handles.append(ax.bar(positions[:len(negative_values)], negative_values, width, color=colors[i], label=labels[i]))

    ax.set_ylabel('Cumulative', color='k')
    ax.tick_params('y', colors='k', labelsize='small')  # 눈금의 글자 크기 조절

    # 주축 설정
    ax.spines['right'].set_visible(False)  # 오른쪽 축 숨김
    ax.yaxis.tick_left()  # 왼쪽 축만 표시

    # x축 눈금 설정
    ax.set_xticks(positions)
    ax.set_xticklabels(x[:len(y)])  # x 눈금 레이블 설정

    # 범례에서 수치가 0인 항목은 제외
    filtered_handles = [h[0] for h in handles if np.sum(h[0].get_height()) != 0]
    filtered_labels = [labels[i] for i in sorted_indexes if np.sum(positive_data.iloc[:, i]) != 0]
    ax.legend(filtered_handles, filtered_labels)
    
    plt.show()

    

################################################## 공통조건 설정 ##################################################

principles_frame = tk.Frame(window, relief="solid", bd=1)
principles_frame.grid(row=2, column=0, columnspan=4, padx=5, pady=5)

principles_label = tk.Label(principles_frame, text="주요 전제조건")
principles_label.grid(row=2, column=0, padx=5, pady=5)

target_year_label = tk.Label(principles_frame, text="분석년도")
target_year_label.grid(row=3, column=0, padx=5, pady=5)
target_year_entry = tk.Entry(principles_frame)
target_year_entry.grid(row=3, column=1, padx=5, pady=5)

target_goal_label = tk.Label(principles_frame, text="감축목표")
target_goal_label.grid(row=4, column=0, padx=5, pady=5)
target_goal_entry = tk.Entry(principles_frame)
target_goal_entry.grid(row=4, column=1, padx=5, pady=5)


energy_demand_change_label = tk.Label(principles_frame, text="에너지 수요변화(%/year)")
energy_demand_change_label.grid(row=5, column=0, padx=5, pady=5)
energy_demand_change_entry = tk.Entry(principles_frame)
energy_demand_change_entry.grid(row=5, column=1, padx=5, pady=5)

energy_self_label = tk.Label(principles_frame, text="에너지 자급률(%)")
energy_self_label.grid(row=6, column=0, padx=5, pady=5)
energy_self_entry = tk.Entry(principles_frame)
energy_self_entry.grid(row=6, column=1, padx=5, pady=5)

land_utilisation_label = tk.Label(principles_frame, text="유휴부지 이용률(%)")
land_utilisation_label.grid(row=7, column=0, padx=5, pady=5)
land_utilisation_entry = tk.Entry(principles_frame)
land_utilisation_entry.grid(row=7, column=1, padx=5, pady=5)


################################################## 유형선택 ##################################################

region_frame = tk.Frame(window, relief="solid", bd=1)
region_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)

region_label = tk.Label(region_frame, text="유형 선택")
region_label.grid(row=0, column=0, padx=5, pady=5)

selected_regions = []

def toggle_region(region):
    if region in selected_regions:
        selected_regions.remove(region)
    else:
        selected_regions.append(region)
    toggle_scenario()

regions = ["수요도시형", "산업중심형", "농어촌지역형", "항만인프라형"]


for i, region in enumerate(regions):
    var = tk.StringVar()
    region_checkbutton = tk.Checkbutton(region_frame, text=region, variable=var, onvalue=region, offvalue="", bd=1, command=lambda region=region: toggle_region(region))
    region_checkbutton.grid(row=1, column=i+1, padx=5, pady=5)
    region_checkbutton.config(width=15)

def create_checkbox_group(parent, items):
    checkboxes = []
    for i, item in enumerate(items):
        var = tk.StringVar()
        checkbox = tk.Checkbutton(parent, text=item, variable=var, onvalue=item, offvalue="", bd=1, command=lambda item=item: toggle_region(item))
        checkbox.grid(row=1, column=i+1, padx=5, pady=5)
        checkbox.config(width=15)
        checkboxes.append(checkbox)
    return checkboxes

def toggle_region(region):
    if region in selected_regions:
        selected_regions.remove(region)
    else:
        selected_regions.append(region)
        # 나머지 체크박스 비활성화
        for checkbox in checkboxes:
            if checkbox["text"] != region:
                checkbox.config(state="disabled")
    # 선택해제 시 모든 체크박스 활성화
    if not selected_regions:
        for checkbox in checkboxes:
            checkbox.config(state="normal")
    toggle_scenario()

# 체크박스 그룹 생성
checkboxes = create_checkbox_group(region_frame, regions)


################################################## 수요도시형 ##################################################

scenario_frame_city = tk.Frame(window, relief="solid", bd=1)
scenario_frame_city.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_city.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_city = tk.Label(scenario_frame_city, text="에너지수요 설정 - 수요도시형")
energy_demand_setting_label_city.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_city, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

city_area_label = tk.Label(scenario_frame_city, text="지역면적(km2)")
city_area_label.grid(row=2, column=0, padx=5, pady=5)
city_area_entry = tk.Entry(scenario_frame_city)
city_area_entry.grid(row=2, column=1, padx=5, pady=5)

residential_area_label = tk.Label(scenario_frame_city, text="주거지역(%)")
residential_area_label.grid(row=3, column=0, padx=5, pady=5)
residential_area_entry = tk.Entry(scenario_frame_city)
residential_area_entry.grid(row=3, column=1, padx=5, pady=5)

commercial_area_label = tk.Label(scenario_frame_city, text="상업지역(%)")
commercial_area_label.grid(row=4, column=0, padx=5, pady=5)
commercial_area_entry = tk.Entry(scenario_frame_city)
commercial_area_entry.grid(row=4, column=1, padx=5, pady=5)

transportation_facility_label = tk.Label(scenario_frame_city, text="교통시설(%)")
transportation_facility_label.grid(row=5, column=0, padx=5, pady=5)
transportation_facility_entry = tk.Entry(scenario_frame_city)
transportation_facility_entry.grid(row=5, column=1, padx=5, pady=5)

environmental_facility_label = tk.Label(scenario_frame_city, text="환경시설(%)")
environmental_facility_label.grid(row=6, column=0, padx=5, pady=5)
environmental_facility_entry = tk.Entry(scenario_frame_city)
environmental_facility_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_city, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

population_change_label = tk.Label(scenario_frame_city, text="인구변화(%/year)")
population_change_label.grid(row=9, column=0, padx=5, pady=5)
population_change_entry = tk.Entry(scenario_frame_city)
population_change_entry.grid(row=9, column=1, padx=5, pady=5)

waste_utilisation_label = tk.Label(scenario_frame_city, text="폐자원 확대(%)")
waste_utilisation_label.grid(row=10, column=0, padx=5, pady=5)
waste_utilisation_entry = tk.Entry(scenario_frame_city)
waste_utilisation_entry.grid(row=10, column=1, padx=5, pady=5)

fes_label = tk.Label(scenario_frame_city, text="최종에너지 구성")
fes_label.grid(row=11, column=0, padx=5, pady=5, sticky="w")

electricity_share_city_label = tk.Label(scenario_frame_city, text="전력화 비중(%)")
electricity_share_city_label.grid(row=12, column=0, padx=5, pady=5)
electricity_share_city_entry = tk.Entry(scenario_frame_city)
electricity_share_city_entry.grid(row=12, column=1, padx=5, pady=5)

h2_share_city_label = tk.Label(scenario_frame_city, text="수소화 비중(%)")
h2_share_city_label.grid(row=13, column=0, padx=5, pady=5)
h2_share_city_entry = tk.Entry(scenario_frame_city)
h2_share_city_entry.grid(row=13, column=1, padx=5, pady=5)


################################################## 산업중심형 ##################################################

scenario_frame_industry = tk.Frame(window, relief="solid", bd=1)
scenario_frame_industry.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_industry.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_industry, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_industry, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_industry, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_industry)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_industry, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_industry)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_industry, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_industry)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)

petro_area_label = tk.Label(scenario_frame_industry, text="석유화학산업(%)")
petro_area_label.grid(row=5, column=0, padx=5, pady=5)
petro_area_entry = tk.Entry(scenario_frame_industry)
petro_area_entry.grid(row=5, column=1, padx=5, pady=5)

cement_area_label = tk.Label(scenario_frame_industry, text="시멘트산업(%)")
cement_area_label.grid(row=6, column=0, padx=5, pady=5)
cement_area_entry = tk.Entry(scenario_frame_industry)
cement_area_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_industry, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_industry, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_industry)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_industry, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_industry)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_industry, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_industry)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)

h2_utilisation_label = tk.Label(scenario_frame_industry, text="[석유]수소대체(%)")
h2_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
h2_utilisation_entry = tk.Entry(scenario_frame_industry)
h2_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)

fuel_eff_label = tk.Label(scenario_frame_industry, text="[산업]에너지효율 향상(%/year)")
fuel_eff_label.grid(row=13, column=0, padx=5, pady=5)
fuel_eff_entry = tk.Entry(scenario_frame_industry)
fuel_eff_entry.grid(row=13, column=1, padx=5, pady=5)

electricity_share_label = tk.Label(scenario_frame_industry, text="[산업]전력화비중 확대(%/year)")
electricity_share_label.grid(row=14, column=0, padx=5, pady=5)
electricity_share_entry = tk.Entry(scenario_frame_industry)
electricity_share_entry.grid(row=14, column=1, padx=5, pady=5)


################################################## 농어촌지역형 ##################################################

scenario_frame_rural = tk.Frame(window, relief="solid", bd=1)
scenario_frame_rural.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_rural.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_rural = tk.Label(scenario_frame_rural, text="에너지수요 설정 - 농어촌지역형")
energy_demand_setting_label_rural.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_rural, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

rural_area_label = tk.Label(scenario_frame_rural, text="지역면적(km2)")
rural_area_label.grid(row=2, column=0, padx=5, pady=5)
rural_area_entry = tk.Entry(scenario_frame_rural)
rural_area_entry.grid(row=2, column=1, padx=5, pady=5)

land_rural_area_label = tk.Label(scenario_frame_rural, text="육상비중(%)")
land_rural_area_label.grid(row=3, column=0, padx=5, pady=5)
land_rural_area_entry = tk.Entry(scenario_frame_rural)
land_rural_area_entry.grid(row=3, column=1, padx=5, pady=5)

sea_rural_area_label = tk.Label(scenario_frame_rural, text="해상비중(%)")
sea_rural_area_label.grid(row=4, column=0, padx=5, pady=5)
sea_rural_area_entry = tk.Entry(scenario_frame_rural)
sea_rural_area_entry.grid(row=4, column=1, padx=5, pady=5)

cultivated_area_label = tk.Label(scenario_frame_rural, text="경작지(km2)")
cultivated_area_label.grid(row=5, column=0, padx=5, pady=5)
cultivated_area_entry = tk.Entry(scenario_frame_rural)
cultivated_area_entry.grid(row=5, column=1, padx=5, pady=5)

forest_label = tk.Label(scenario_frame_rural, text="산림(km2)")
forest_label.grid(row=6, column=0, padx=5, pady=5)
forest_entry = tk.Entry(scenario_frame_rural)
forest_entry.grid(row=6, column=1, padx=5, pady=5)

fish_farm_label = tk.Label(scenario_frame_rural, text="양식장(km2)")
fish_farm_label.grid(row=7, column=0, padx=5, pady=5)
fish_farm_entry = tk.Entry(scenario_frame_rural)
fish_farm_entry.grid(row=7, column=1, padx=5, pady=5)

fishing_boat_label = tk.Label(scenario_frame_rural, text="어선(대)")
fishing_boat_label.grid(row=8, column=0, padx=5, pady=5)
fishing_boat_entry = tk.Entry(scenario_frame_rural)
fishing_boat_entry.grid(row=8, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_rural, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=10, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

trees_label = tk.Label(scenario_frame_rural, text="[농촌]임산변화율(%)")
trees_label.grid(row=11, column=0, padx=5, pady=5)
trees_entry = tk.Entry(scenario_frame_rural)
trees_entry.grid(row=11, column=1, padx=5, pady=5)

rice_label = tk.Label(scenario_frame_rural, text="[농촌]농산변화율(%)")
rice_label.grid(row=12, column=0, padx=5, pady=5)
rice_entry = tk.Entry(scenario_frame_rural)
rice_entry.grid(row=12, column=1, padx=5, pady=5)

animals_label = tk.Label(scenario_frame_rural, text="[농촌]축산변화율(%)")
animals_label.grid(row=13, column=0, padx=5, pady=5)
animals_entry = tk.Entry(scenario_frame_rural)
animals_entry.grid(row=13, column=1, padx=5, pady=5)

fishfarm_change_label = tk.Label(scenario_frame_rural, text="[어촌]양식장 면적증가율(%/y)")
fishfarm_change_label.grid(row=14, column=0, padx=5, pady=5)
fishfarm_change_entry = tk.Entry(scenario_frame_rural)
fishfarm_change_entry.grid(row=14, column=1, padx=5, pady=5)

electric_boat_label = tk.Label(scenario_frame_rural, text="[어촌]선박탈탄소(전력화)(%)")
electric_boat_label.grid(row=15, column=0, padx=5, pady=5)
electric_boat_entry = tk.Entry(scenario_frame_rural)
electric_boat_entry.grid(row=15, column=1, padx=5, pady=5)

h2_boat_label = tk.Label(scenario_frame_rural, text="[어촌]선박탈탄소(수소화)(%)")
h2_boat_label.grid(row=16, column=0, padx=5, pady=5)
h2_boat_entry = tk.Entry(scenario_frame_rural)
h2_boat_entry.grid(row=16, column=1, padx=5, pady=5)

P2G_share_label = tk.Label(scenario_frame_rural, text="잉여전력의 수소생산 비중(%)")
P2G_share_label.grid(row=17, column=0, padx=5, pady=5)
P2G_share_entry = tk.Entry(scenario_frame_rural)
P2G_share_entry.grid(row=17, column=1, padx=5, pady=5)


################################################## 항만인프라형 ##################################################

scenario_frame_port = tk.Frame(window, relief="solid", bd=1)
scenario_frame_port.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_port.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_port = tk.Label(scenario_frame_port, text="에너지수요 설정 - 항만인프라형")
energy_demand_setting_label_port.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_port, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

port_area_label = tk.Label(scenario_frame_port, text="항만면적(km2)")
port_area_label.grid(row=2, column=0, padx=5, pady=5)
port_area_entry = tk.Entry(scenario_frame_port)
port_area_entry.grid(row=2, column=1, padx=5, pady=5)

sea_port_area_label = tk.Label(scenario_frame_port, text="해역면적(km2)")
sea_port_area_label.grid(row=3, column=0, padx=5, pady=5)
sea_port_area_entry = tk.Entry(scenario_frame_port)
sea_port_area_entry.grid(row=3, column=1, padx=5, pady=5)

cargo_cap_label = tk.Label(scenario_frame_port, text="화물처리(톤)")
cargo_cap_label.grid(row=4, column=0, padx=5, pady=5)
cargo_cap_entry = tk.Entry(scenario_frame_port)
cargo_cap_entry.grid(row=4, column=1, padx=5, pady=5)



# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_port, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=5, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

petro_share_label = tk.Label(scenario_frame_port, text="석유물동량 처리비중(%)")
petro_share_label.grid(row=6, column=0, padx=5, pady=5)
petro_share_entry = tk.Entry(scenario_frame_port)
petro_share_entry.grid(row=6, column=1, padx=5, pady=5)

trans_h2_label = tk.Label(scenario_frame_port, text="석유->수소 대체비중(%)")
trans_h2_label.grid(row=7, column=0, padx=5, pady=5)
trans_h2_entry = tk.Entry(scenario_frame_port)
trans_h2_entry.grid(row=7, column=1, padx=5, pady=5)

import_h2_label = tk.Label(scenario_frame_port, text="수소수입형태")
import_h2_label.grid(row=8, column=0, padx=5, pady=5, sticky="w")

liquid_h2_label = tk.Label(scenario_frame_port, text="액화수소(%)")
liquid_h2_label.grid(row=9, column=0, padx=5, pady=5)
liquid_h2_entry = tk.Entry(scenario_frame_port)
liquid_h2_entry.grid(row=9, column=1, padx=5, pady=5)

amonia_label = tk.Label(scenario_frame_port, text="암모니아(%)")
amonia_label.grid(row=10, column=0, padx=5, pady=5)
amonia_entry = tk.Entry(scenario_frame_port)
amonia_entry.grid(row=10, column=1, padx=5, pady=5)

storage_cap_label = tk.Label(scenario_frame_port, text="비축분(일분)")
storage_cap_label.grid(row=11, column=0, padx=5, pady=5)
storage_cap_entry = tk.Entry(scenario_frame_port)
storage_cap_entry.grid(row=11, column=1, padx=5, pady=5)

wind_farm_label = tk.Label(scenario_frame_port, text="해상풍력 이용")
wind_farm_label.grid(row=12, column=0, padx=5, pady=5, sticky="w")

sea_area_change_label = tk.Label(scenario_frame_port, text="이용가능 해역면적 변화(%/y)")
sea_area_change_label.grid(row=13, column=0, padx=5, pady=5)
sea_area_change_entry = tk.Entry(scenario_frame_port)
sea_area_change_entry.grid(row=13, column=1, padx=5, pady=5)

density_change_label = tk.Label(scenario_frame_port, text="풍력단지 밀집도 변화(%/y)")
density_change_label.grid(row=14, column=0, padx=5, pady=5)
density_change_entry = tk.Entry(scenario_frame_port)
density_change_entry.grid(row=14, column=1, padx=5, pady=5)

dedicated_h2_label = tk.Label(scenario_frame_port, text="수소 고정생산 비중(%)")
dedicated_h2_label.grid(row=15, column=0, padx=5, pady=5)
dedicated_h2_entry = tk.Entry(scenario_frame_port)
dedicated_h2_entry.grid(row=15, column=1, padx=5, pady=5)



################################################## 저장 및 실행 ##################################################

save_button = tk.Button(window, text="Save to Excel", command=save_to_excel)
save_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="w")

run_button = tk.Button(window, text="Run Script", command=run_script)
run_button.grid(row=5, column=2, columnspan=2, padx=5, pady=5, sticky="e")

result_button = tk.Button(window, text="Result", command=show_result)
result_button.grid(row=6, column=0, columnspan=4, padx=5, pady=5)

window.mainloop()

