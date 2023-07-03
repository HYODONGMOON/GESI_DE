# -*- coding: utf-8 -*-

import tkinter as tk
import subprocess
from openpyxl import load_workbook

def toggle_entry(entry, var):
    if var.get():
        entry.config(state='normal')
    else:
        entry.config(state='disabled')

def save_to_excel():
    target_year = target_year_entry.get()
    target_goal = target_goal_entry.get()

    wb = load_workbook('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')
    ws = wb['Scenario']
    ws['E4'] = target_year
    ws['E5'] = target_goal
    wb.save('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')

def run_script():
    subprocess.call(["python", "C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\gesi.py"])

def toggle_scenario():
    if len(selected_regions) >= 2:
        for checkbox in region_checkboxes:
            checkbox.config(state='disabled')
    else:
        for checkbox in region_checkboxes:
            checkbox.config(state='normal')

window = tk.Tk()
window.geometry("700x800")


################################################## 공통조건 설정 ##################################################

principles_frame = tk.Frame(window, relief="solid", bd=1)
principles_frame.grid(row=0, column=0, columnspan=4, padx=5, pady=5)

principles_label = tk.Label(principles_frame, text="주요 전제조건")
principles_label.grid(row=0, column=0, padx=5, pady=5)

target_year_label = tk.Label(principles_frame, text="분석년도")
target_year_label.grid(row=1, column=0, padx=5, pady=5)
target_year_entry = tk.Entry(principles_frame)
target_year_entry.grid(row=1, column=1, padx=5, pady=5)

target_goal_label = tk.Label(principles_frame, text="감축목표")
target_goal_label.grid(row=2, column=0, padx=5, pady=5)
target_goal_entry = tk.Entry(principles_frame)
target_goal_entry.grid(row=2, column=1, padx=5, pady=5)

energy_demand_change_label = tk.Label(principles_frame, text="에너지 수요변화(%/year)")
energy_demand_change_label.grid(row=3, column=0, padx=5, pady=5)
energy_demand_change_entry = tk.Entry(principles_frame)
energy_demand_change_entry.grid(row=3, column=1, padx=5, pady=5)

energy_self_label = tk.Label(principles_frame, text="에너지 자급률(%)")
energy_self_label.grid(row=4, column=0, padx=5, pady=5)
energy_self_entry = tk.Entry(principles_frame)
energy_self_entry.grid(row=4, column=1, padx=5, pady=5)

land_utilisation_label = tk.Label(principles_frame, text="유휴부지 이용률(%)")
land_utilisation_label.grid(row=5, column=0, padx=5, pady=5)
land_utilisation_entry = tk.Entry(principles_frame)
land_utilisation_entry.grid(row=5, column=1, padx=5, pady=5)

################################################## 유형선택 ##################################################

region_frame = tk.Frame(window, relief="solid", bd=1)
region_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)

region_label = tk.Label(region_frame, text="유형 선택")
region_label.grid(row=0, column=0, padx=5, pady=5)

selected_regions = []

def toggle_region(region):
    if region in selected_regions:
        selected_regions.remove(region)
    else:
        if len(selected_regions) >= 2:
            return
        selected_regions.append(region)
    toggle_scenario()
    update_grid_visibility()

def update_grid_visibility():
    for region in regions:
        if region in selected_regions:
            grid_mapping[region].grid(row=4, column=grid_column_mapping[region], columnspan=2, padx=5, pady=5, sticky="w")
        else:
            grid_mapping[region].grid_forget()

def enable_all_checkboxes():
    for checkbox in region_checkboxes:
        checkbox.config(state="normal")    

region_checkboxes = []
grid_mapping = {}
grid_column_mapping = {}


regions = ["수요도시형", "산업중심형", "농어촌지역형", "항만인프라형"]


for i, region in enumerate(regions):
    var = tk.StringVar()
    region_checkbutton = tk.Checkbutton(region_frame, text=region, variable=var, onvalue=region, offvalue="", bd=1, command=lambda region=region: toggle_region(region))
    region_checkbutton.grid(row=1, column=i+1, padx=5, pady=5)
    region_checkbutton.config(width=15)
    region_checkboxes.append(region_checkbutton)

    region_grid_frame = tk.Frame(window, relief="solid", bd=1)
    grid_mapping[region] = region_grid_frame
    grid_column_mapping[region] = i * 2
    region_grid_frame.grid(row=4, column=i * 2, columnspan=2, padx=5, pady=5)
    region_grid_frame.grid_forget()

################################################## 수요도시형 ##################################################

scenario_frame_city = tk.Frame(window, relief="solid", bd=1)
scenario_frame_city.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_city.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_city = tk.Label(scenario_frame_city, text="에너지수요 설정 - 수요도시형")
energy_demand_setting_label_city.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_city, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_city, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_city)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

residential_area_label = tk.Label(scenario_frame_city, text="주거지역(%)")
residential_area_label.grid(row=3, column=0, padx=5, pady=5)
residential_area_entry = tk.Entry(scenario_frame_city)
residential_area_entry.grid(row=3, column=1, padx=5, pady=5)

commercial_area_label = tk.Label(scenario_frame_city, text="상업지역(%)")
commercial_area_label.grid(row=4, column=0, padx=5, pady=5)
commercial_area_entry = tk.Entry(scenario_frame_city)
commercial_area_entry.grid(row=4, column=1, padx=5, pady=5)

transportation_facility_label = tk.Label(scenario_frame_city, text="교통시설(%)")
transportation_facility_label.grid(row=5, column=0, padx=5, pady=5)
transportation_facility_entry = tk.Entry(scenario_frame_city)
transportation_facility_entry.grid(row=5, column=1, padx=5, pady=5)

environmental_facility_label = tk.Label(scenario_frame_city, text="환경시설(%)")
environmental_facility_label.grid(row=6, column=0, padx=5, pady=5)
environmental_facility_entry = tk.Entry(scenario_frame_city)
environmental_facility_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_city, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

population_change_label = tk.Label(scenario_frame_city, text="인구변화(%/year)")
population_change_label.grid(row=9, column=0, padx=5, pady=5)
population_change_entry = tk.Entry(scenario_frame_city)
population_change_entry.grid(row=9, column=1, padx=5, pady=5)

waste_utilisation_label = tk.Label(scenario_frame_city, text="폐자원 확대(%)")
waste_utilisation_label.grid(row=10, column=0, padx=5, pady=5)
waste_utilisation_entry = tk.Entry(scenario_frame_city)
waste_utilisation_entry.grid(row=10, column=1, padx=5, pady=5)

land_utilisation_label = tk.Label(scenario_frame_city, text="최종에너지 구성")
land_utilisation_label.grid(row=11, column=0, padx=5, pady=5, sticky="w")

land_utilisation_label = tk.Label(scenario_frame_city, text="전력화 비중(%)")
land_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
land_utilisation_entry = tk.Entry(scenario_frame_city)
land_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)

land_utilisation_label = tk.Label(scenario_frame_city, text="수소화 비중(%)")
land_utilisation_label.grid(row=13, column=0, padx=5, pady=5)
land_utilisation_entry = tk.Entry(scenario_frame_city)
land_utilisation_entry.grid(row=13, column=1, padx=5, pady=5)


################################################## 산업중심형 ##################################################

scenario_frame_industry = tk.Frame(window, relief="solid", bd=1)
scenario_frame_industry.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_industry.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_industry, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_industry, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_industry, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_industry)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_industry, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_industry)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_industry, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_industry)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)

petro_area_label = tk.Label(scenario_frame_industry, text="석유화학산업(%)")
petro_area_label.grid(row=5, column=0, padx=5, pady=5)
petro_area_entry = tk.Entry(scenario_frame_industry)
petro_area_entry.grid(row=5, column=1, padx=5, pady=5)

cement_area_label = tk.Label(scenario_frame_industry, text="시멘트산업(%)")
cement_area_label.grid(row=6, column=0, padx=5, pady=5)
cement_area_entry = tk.Entry(scenario_frame_industry)
cement_area_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_industry, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_industry, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_industry)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_industry, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_industry)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_industry, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_industry)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)

h2_utilisation_label = tk.Label(scenario_frame_industry, text="[석유]수소대체(%)")
h2_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
h2_utilisation_entry = tk.Entry(scenario_frame_industry)
h2_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)

fuel_eff_label = tk.Label(scenario_frame_industry, text="[산업]에너지효율 향상(%/year)")
fuel_eff_label.grid(row=13, column=0, padx=5, pady=5)
fuel_eff_entry = tk.Entry(scenario_frame_industry)
fuel_eff_entry.grid(row=13, column=1, padx=5, pady=5)

electricity_share_label = tk.Label(scenario_frame_industry, text="[산업]전력화비중 확대(%/year)")
electricity_share_label.grid(row=14, column=0, padx=5, pady=5)
electricity_share_entry = tk.Entry(scenario_frame_industry)
electricity_share_entry.grid(row=14, column=1, padx=5, pady=5)


################################################## 농어촌지역형 ##################################################

scenario_frame_rural = tk.Frame(window, relief="solid", bd=1)
scenario_frame_rural.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_rural.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_rural, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_rural, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_rural, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_rural)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_rural, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_rural)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_rural, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_industry)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)

petro_area_label = tk.Label(scenario_frame_rural, text="석유화학산업(%)")
petro_area_label.grid(row=5, column=0, padx=5, pady=5)
petro_area_entry = tk.Entry(scenario_frame_industry)
petro_area_entry.grid(row=5, column=1, padx=5, pady=5)

cement_area_label = tk.Label(scenario_frame_rural, text="시멘트산업(%)")
cement_area_label.grid(row=6, column=0, padx=5, pady=5)
cement_area_entry = tk.Entry(scenario_frame_industry)
cement_area_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_rural, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_rural, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_rural)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_rural, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_rural)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_rural, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_rural)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)

h2_utilisation_label = tk.Label(scenario_frame_rural, text="[석유]수소대체(%)")
h2_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
h2_utilisation_entry = tk.Entry(scenario_frame_rural)
h2_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)


################################################## 항만인프라형 ##################################################

scenario_frame_port = tk.Frame(window, relief="solid", bd=1)
scenario_frame_port.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_port.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_port, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_port, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_port, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_port)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_port, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_port)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_port, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_port)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_port, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_port, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_port)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_port, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_port)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_port, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_port)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)


################################################## 저장 및 실행 ##################################################

save_button = tk.Button(window, text="Save to Excel", command=save_to_excel)
save_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="w")

run_button = tk.Button(window, text="Run Script", command=run_script)
run_button.grid(row=5, column=2, columnspan=2, padx=5, pady=5, sticky="e")

window.mainloop()