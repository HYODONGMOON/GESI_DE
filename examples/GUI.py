# -*- coding: utf-8 -*-

import tkinter as tk
import subprocess
import tkinter.messagebox as messagebox
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from tkinter import ttk


# 전체 인터페이스 코드 #

def handle_menu_selection(event):
    selected_item = menu_var.get()
    print(f"Selected: {selected_item}")

window = tk.Tk()
# window.geometry("700x800")

# 최상단 메뉴 패널
menubar = tk.Menu(window)
window.config(menu=menubar)

# "파일" 메뉴
file_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="파일", menu=file_menu)
file_menu.add_command(label="열기")
file_menu.add_command(label="저장")
file_menu.add_separator()
file_menu.add_command(label="종료")

# "편집" 메뉴
edit_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="편집", menu=edit_menu)
edit_menu.add_command(label="복사")
edit_menu.add_command(label="붙여넣기")
edit_menu.add_separator()
edit_menu.add_command(label="실행 취소")

# "도움말" 메뉴
help_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label="도움말", menu=help_menu)
help_menu.add_command(label="도움말 보기")
help_menu.add_command(label="버전 정보")

# 나머지 인터페이스 그리드
# ...


# 모델 분석관련 코드 #

def toggle_entry(entry, var):
    if var.get():
        entry.config(state='normal')
    else:
        entry.config(state='disabled')

# 입력값 제한

def save_to_excel():
    input_fields = {
        "감축목표": {
            "entry": target_goal_entry, 
            "min_val": 10, 
            "max_val": 100
            },
        "에너지 자급률": {
            "entry": energy_self_entry,
            "min_val": 0, 
            "max_val": 50
            },
        "인구 변화": {
            "entry": population_change_entry, 
            "min_val": 0, 
            "max_val": 50
            } 
        }

    error_messages = []

    for field_name, field_info in input_fields.items():
        entry = field_info["entry"]
        value = entry.get()
        min_val = field_info["min_val"]
        max_val = field_info["max_val"]

        if not validate_input(value, min_val, max_val):
            error_messages.append(field_name)
            entry.config(bg="red")  # 잘못된 입력 필드에 빨간색 음영 적용
        else:
            entry.config(bg="white")  # 올바른 입력 필드는 흰색으로 설정

    if error_messages:
        show_error_message(error_messages)
        return

    target_year = int(target_year_entry.get())
    target_goal = float(target_goal_entry.get())
    energy_demand_change = float(energy_demand_change_entry.get())
    energy_self = float(energy_self_entry.get())
    land_utilisation = float(land_utilisation_entry.get())
    population_change = float(population_change_entry.get())
    electricity_share_city = float(electricity_share_city_entry.get())
    h2_share_city = float(h2_share_city_entry.get())
    waste_utilisation = float(waste_utilisation_entry.get())
    
    city_area = float(city_area_entry.get())
    residential_area = float(residential_area_entry.get())
    commercial_area = float(commercial_area_entry.get())
    transportation_facility = float(transportation_facility_entry.get())
    environmental_facility = float(environmental_facility_entry.get())


    wb = load_workbook('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')
    ws = wb['Scenario']
    
    # 기본 전제조건
    ws['E4'] = target_year
    ws['E5'] = target_goal /100
    ws['E7'] = energy_demand_change /100
    ws['E11'] = energy_self /100
    ws['E16'] = land_utilisation /100

    #수요도시형 시나리오
    ws['E22'] = population_change /100
    ws['E23'] = population_change /100
    ws['E25'] = electricity_share_city /100
    ws['E28'] = electricity_share_city /100
    ws['E27'] = h2_share_city /100
    ws['E29'] = h2_share_city /100
    ws['E31'] = waste_utilisation /100

    #산업중심형 시나리오


    ws1 = wb['Demand']

    #수요도시형 
    ws1['E4'] = city_area
    ws1['E5'] = residential_area /100
    ws1['E6'] = commercial_area /100
    ws1['E7'] = transportation_facility /100
    ws1['E8'] = environmental_facility /100

    #산업중심형


    wb.save('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\GESI_DE.xlsx')

   

def run_script():
    subprocess.call(["python", "C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\gesi.py"])

def toggle_scenario():
    if all(region in selected_regions for region in regions):
        scenario_frame_city.grid(row=4, column=0, columnspan=2, padx=5, pady=5, sticky="w")
        scenario_frame_industry.grid(row=4, column=2, columnspan=2, padx=5, pady=5, sticky="e")
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "수요도시형" in selected_regions:
        scenario_frame_city.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "산업중심형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()
    elif "농어촌지역형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
        scenario_frame_port.grid_forget()
    elif "항만인프라형" in selected_regions:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
    else:
        scenario_frame_city.grid_forget()
        scenario_frame_industry.grid_forget()
        scenario_frame_rural.grid_forget()
        scenario_frame_port.grid_forget()

def validate_input(value, min_val, max_val):
    try:
        num = float(value)
        if min_val <= num <= max_val:
            return True
        else:
            return False
    except ValueError:
        return False

def show_error_message(error_messages):
    error_text = "\n".join(f" - {field_name}" for field_name in error_messages)
    messagebox.showerror("입력 오류", f"다음 입력 필드가 올바른 범위 내에 입력되지 않았습니다:\n{error_text}")


# 데이터 시각화 #


def show_result():
    # 엑셀 파일에서 데이터 읽기
    df = pd.read_excel('C:\\Users\\GESI Moon\\Desktop\\python workplace\\GESI_DE - GUI\\examples\\results_2050.xlsx', sheet_name='Figure_hour_electricity', header=None, usecols="W:AO", nrows=72)

    # 데이터 출력
    print(df)

    # 데이터 시각화
    fig, ax1 = plt.subplots()

    # X행 선 그래프
    x = [str(i) for i in df.index]  # 정수형 인덱스를 문자열로 변환
    y = df.iloc[:, 0]
    ax1.plot(x, y, 'b-')
    ax1.set_xlabel('X')
    ax1.set_ylabel('Y', color='b')
    ax1.tick_params('y', colors='b')

    ax2 = ax1.twinx()

    # Y~AO행 누적 막대 그래프
    cumulative_data = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').cumsum(axis=1)  # 문자열을 부동소수점으로 변환 후 누적 합계 계산
    colors = ['r', 'g', 'c', 'm', 'y'] * (len(cumulative_data.columns) // 5 + 1)  # 막대 그래프 색상
    for i, col in enumerate(cumulative_data.columns):
        ax2.bar(x, cumulative_data[col], color=colors[i], label=col)

    ax2.set_ylabel('Cumulative', color='k')
    ax2.tick_params('y', colors='k')

    plt.legend()
    plt.show()

    

################################################## 공통조건 설정 ##################################################

principles_frame = tk.Frame(window, relief="solid", bd=1)
principles_frame.grid(row=2, column=0, columnspan=4, padx=5, pady=5)

principles_label = tk.Label(principles_frame, text="주요 전제조건")
principles_label.grid(row=2, column=0, padx=5, pady=5)

target_year_label = tk.Label(principles_frame, text="분석년도")
target_year_label.grid(row=3, column=0, padx=5, pady=5)
target_year_entry = tk.Entry(principles_frame)
target_year_entry.grid(row=3, column=1, padx=5, pady=5)

target_goal_label = tk.Label(principles_frame, text="감축목표")
target_goal_label.grid(row=4, column=0, padx=5, pady=5)
target_goal_entry = tk.Entry(principles_frame)
target_goal_entry.grid(row=4, column=1, padx=5, pady=5)


energy_demand_change_label = tk.Label(principles_frame, text="에너지 수요변화(%/year)")
energy_demand_change_label.grid(row=5, column=0, padx=5, pady=5)
energy_demand_change_entry = tk.Entry(principles_frame)
energy_demand_change_entry.grid(row=5, column=1, padx=5, pady=5)

energy_self_label = tk.Label(principles_frame, text="에너지 자급률(%)")
energy_self_label.grid(row=6, column=0, padx=5, pady=5)
energy_self_entry = tk.Entry(principles_frame)
energy_self_entry.grid(row=6, column=1, padx=5, pady=5)

land_utilisation_label = tk.Label(principles_frame, text="유휴부지 이용률(%)")
land_utilisation_label.grid(row=7, column=0, padx=5, pady=5)
land_utilisation_entry = tk.Entry(principles_frame)
land_utilisation_entry.grid(row=7, column=1, padx=5, pady=5)

################################################## 유형선택 ##################################################

region_frame = tk.Frame(window, relief="solid", bd=1)
region_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)

region_label = tk.Label(region_frame, text="유형 선택")
region_label.grid(row=0, column=0, padx=5, pady=5)

selected_regions = []

def toggle_region(region):
    if region in selected_regions:
        selected_regions.remove(region)
    else:
        selected_regions.append(region)
    toggle_scenario()

regions = ["수요도시형", "산업중심형", "농어촌지역형", "항만인프라형"]


for i, region in enumerate(regions):
    var = tk.StringVar()
    region_checkbutton = tk.Checkbutton(region_frame, text=region, variable=var, onvalue=region, offvalue="", bd=1, command=lambda region=region: toggle_region(region))
    region_checkbutton.grid(row=1, column=i+1, padx=5, pady=5)
    region_checkbutton.config(width=15)

################################################## 수요도시형 ##################################################

scenario_frame_city = tk.Frame(window, relief="solid", bd=1)
scenario_frame_city.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_city.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_city = tk.Label(scenario_frame_city, text="에너지수요 설정 - 수요도시형")
energy_demand_setting_label_city.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_city, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

city_area_label = tk.Label(scenario_frame_city, text="지역면적(km2)")
city_area_label.grid(row=2, column=0, padx=5, pady=5)
city_area_entry = tk.Entry(scenario_frame_city)
city_area_entry.grid(row=2, column=1, padx=5, pady=5)

residential_area_label = tk.Label(scenario_frame_city, text="주거지역(%)")
residential_area_label.grid(row=3, column=0, padx=5, pady=5)
residential_area_entry = tk.Entry(scenario_frame_city)
residential_area_entry.grid(row=3, column=1, padx=5, pady=5)

commercial_area_label = tk.Label(scenario_frame_city, text="상업지역(%)")
commercial_area_label.grid(row=4, column=0, padx=5, pady=5)
commercial_area_entry = tk.Entry(scenario_frame_city)
commercial_area_entry.grid(row=4, column=1, padx=5, pady=5)

transportation_facility_label = tk.Label(scenario_frame_city, text="교통시설(%)")
transportation_facility_label.grid(row=5, column=0, padx=5, pady=5)
transportation_facility_entry = tk.Entry(scenario_frame_city)
transportation_facility_entry.grid(row=5, column=1, padx=5, pady=5)

environmental_facility_label = tk.Label(scenario_frame_city, text="환경시설(%)")
environmental_facility_label.grid(row=6, column=0, padx=5, pady=5)
environmental_facility_entry = tk.Entry(scenario_frame_city)
environmental_facility_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_city, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

population_change_label = tk.Label(scenario_frame_city, text="인구변화(%/year)")
population_change_label.grid(row=9, column=0, padx=5, pady=5)
population_change_entry = tk.Entry(scenario_frame_city)
population_change_entry.grid(row=9, column=1, padx=5, pady=5)

waste_utilisation_label = tk.Label(scenario_frame_city, text="폐자원 확대(%)")
waste_utilisation_label.grid(row=10, column=0, padx=5, pady=5)
waste_utilisation_entry = tk.Entry(scenario_frame_city)
waste_utilisation_entry.grid(row=10, column=1, padx=5, pady=5)

fes_label = tk.Label(scenario_frame_city, text="최종에너지 구성")
fes_label.grid(row=11, column=0, padx=5, pady=5, sticky="w")

electricity_share_city_label = tk.Label(scenario_frame_city, text="전력화 비중(%)")
electricity_share_city_label.grid(row=12, column=0, padx=5, pady=5)
electricity_share_city_entry = tk.Entry(scenario_frame_city)
electricity_share_city_entry.grid(row=12, column=1, padx=5, pady=5)

h2_share_city_label = tk.Label(scenario_frame_city, text="수소화 비중(%)")
h2_share_city_label.grid(row=13, column=0, padx=5, pady=5)
h2_share_city_entry = tk.Entry(scenario_frame_city)
h2_share_city_entry.grid(row=13, column=1, padx=5, pady=5)


################################################## 산업중심형 ##################################################

scenario_frame_industry = tk.Frame(window, relief="solid", bd=1)
scenario_frame_industry.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_industry.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_industry, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_industry, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_industry, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_industry)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_industry, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_industry)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_industry, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_industry)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)

petro_area_label = tk.Label(scenario_frame_industry, text="석유화학산업(%)")
petro_area_label.grid(row=5, column=0, padx=5, pady=5)
petro_area_entry = tk.Entry(scenario_frame_industry)
petro_area_entry.grid(row=5, column=1, padx=5, pady=5)

cement_area_label = tk.Label(scenario_frame_industry, text="시멘트산업(%)")
cement_area_label.grid(row=6, column=0, padx=5, pady=5)
cement_area_entry = tk.Entry(scenario_frame_industry)
cement_area_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_industry, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_industry, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_industry)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_industry, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_industry)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_industry, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_industry)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)

h2_utilisation_label = tk.Label(scenario_frame_industry, text="[석유]수소대체(%)")
h2_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
h2_utilisation_entry = tk.Entry(scenario_frame_industry)
h2_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)

fuel_eff_label = tk.Label(scenario_frame_industry, text="[산업]에너지효율 향상(%/year)")
fuel_eff_label.grid(row=13, column=0, padx=5, pady=5)
fuel_eff_entry = tk.Entry(scenario_frame_industry)
fuel_eff_entry.grid(row=13, column=1, padx=5, pady=5)

electricity_share_label = tk.Label(scenario_frame_industry, text="[산업]전력화비중 확대(%/year)")
electricity_share_label.grid(row=14, column=0, padx=5, pady=5)
electricity_share_entry = tk.Entry(scenario_frame_industry)
electricity_share_entry.grid(row=14, column=1, padx=5, pady=5)


################################################## 농어촌지역형 ##################################################

scenario_frame_rural = tk.Frame(window, relief="solid", bd=1)
scenario_frame_rural.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_rural.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_rural, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_rural, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_rural, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_rural)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_rural, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_rural)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_rural, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_industry)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)

petro_area_label = tk.Label(scenario_frame_rural, text="석유화학산업(%)")
petro_area_label.grid(row=5, column=0, padx=5, pady=5)
petro_area_entry = tk.Entry(scenario_frame_industry)
petro_area_entry.grid(row=5, column=1, padx=5, pady=5)

cement_area_label = tk.Label(scenario_frame_rural, text="시멘트산업(%)")
cement_area_label.grid(row=6, column=0, padx=5, pady=5)
cement_area_entry = tk.Entry(scenario_frame_industry)
cement_area_entry.grid(row=6, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_rural, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_rural, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_rural)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_rural, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_rural)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_rural, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_rural)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)

h2_utilisation_label = tk.Label(scenario_frame_rural, text="[석유]수소대체(%)")
h2_utilisation_label.grid(row=12, column=0, padx=5, pady=5)
h2_utilisation_entry = tk.Entry(scenario_frame_rural)
h2_utilisation_entry.grid(row=12, column=1, padx=5, pady=5)


################################################## 항만인프라형 ##################################################

scenario_frame_port = tk.Frame(window, relief="solid", bd=1)
scenario_frame_port.grid(row=4, column=0, columnspan=4, padx=5, pady=5)
scenario_frame_port.grid_forget()


# 에너지수요 설정 #

energy_demand_setting_label_industry = tk.Label(scenario_frame_port, text="에너지수요 설정 - 산업중심형")
energy_demand_setting_label_industry.grid(row=0, column=0, padx=5, pady=5, columnspan=4)

default_setting_label = tk.Label(scenario_frame_port, text="1. 기본 설정")
default_setting_label.grid(row=1, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

regional_area_label = tk.Label(scenario_frame_port, text="지역면적(km2)")
regional_area_label.grid(row=2, column=0, padx=5, pady=5)
regional_area_entry = tk.Entry(scenario_frame_port)
regional_area_entry.grid(row=2, column=1, padx=5, pady=5)

iron_area_label = tk.Label(scenario_frame_port, text="철강산업(%)")
iron_area_label.grid(row=3, column=0, padx=5, pady=5)
iron_area_entry = tk.Entry(scenario_frame_port)
iron_area_entry.grid(row=3, column=1, padx=5, pady=5)

refinery_area_label = tk.Label(scenario_frame_port, text="정유산업(%)")
refinery_area_label.grid(row=4, column=0, padx=5, pady=5)
refinery_area_entry = tk.Entry(scenario_frame_port)
refinery_area_entry.grid(row=4, column=1, padx=5, pady=5)


# 시나리오 분석가정 입력 #

scenario_setting_label = tk.Label(scenario_frame_port, text="\n2. 시나리오 설정")
scenario_setting_label.grid(row=8, column=0, padx=5, pady=0, sticky="w")  # sticky="w"를 추가하여 왼쪽 정렬

eaf_label = tk.Label(scenario_frame_port, text="[철강]전기로 비중(%)")
eaf_label.grid(row=9, column=0, padx=5, pady=5)
eaf_entry = tk.Entry(scenario_frame_port)
eaf_entry.grid(row=9, column=1, padx=5, pady=5)

dri_label = tk.Label(scenario_frame_port, text="[철강]DRI 비중(%)")
dri_label.grid(row=10, column=0, padx=5, pady=5)
dri_entry = tk.Entry(scenario_frame_port)
dri_entry.grid(row=10, column=1, padx=5, pady=5)

chemical_re_label = tk.Label(scenario_frame_port, text="[석유]화학적재활용(%)")
chemical_re_label.grid(row=11, column=0, padx=5, pady=5)
chemical_re_entry = tk.Entry(scenario_frame_port)
chemical_re_entry.grid(row=11, column=1, padx=5, pady=5)



################################################## 저장 및 실행 ##################################################

save_button = tk.Button(window, text="Save to Excel", command=save_to_excel)
save_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="w")

run_button = tk.Button(window, text="Run Script", command=run_script)
run_button.grid(row=5, column=2, columnspan=2, padx=5, pady=5, sticky="e")

result_button = tk.Button(window, text="Result", command=show_result)
result_button.grid(row=6, column=0, columnspan=4, padx=5, pady=5)

window.mainloop()

